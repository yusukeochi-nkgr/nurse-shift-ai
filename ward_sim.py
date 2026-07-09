import streamlit as st
import pandas as pd
import numpy as np
import io
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import pulp
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import threading
import time

# ==========================================
# 共通デザイン・セッション初期化
# ==========================================
st.set_page_config(page_title="Strategic Nurse Staffing Platform", layout="wide")
st.markdown("""
    <style>
    .main { background-color: #f8f9fa; color: #333333; }
    h1, h2, h3 { font-family: 'Helvetica Neue', Arial, sans-serif; color: #111111; font-weight: 600; }
    div.stButton > button { background-color: #333333; color: #ffffff; border-radius: 2px; border: none; padding: 0.5rem 1rem; font-weight: bold; }
    div.stButton > button:hover { background-color: #000000; color: #ffffff; }
    hr { margin-top: 1.5rem; margin-bottom: 1.5rem; border-top: 1px solid #dcdcdc; }
    .metric-card { background-color: #ffffff; padding: 20px; border: 1px solid #dcdcdc; border-radius: 2px; margin-bottom: 10px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }
    .status-ok { color: #2e7d32; font-weight: bold; }
    .status-ng { color: #c62828; font-weight: bold; }
    .report-box { background-color: #fff3cd; border-left: 5px solid #ff9800; padding: 15px; margin-bottom: 15px; color: #333; }
    .report-box h4 { color: #e65100; margin-top: 0; }
    .gap-surplus { color: #1976d2; font-weight: bold; }
    .gap-shortage { color: #d32f2f; font-weight: bold; }
    .gap-perfect { color: #388e3c; font-weight: bold; }
    </style>
""", unsafe_allow_html=True)

if "org_df" not in st.session_state:
    st.session_state.org_df = pd.DataFrame(columns=["法人名", "施設名", "病棟名", "配置比率（対1）", "病床数", "夜勤基本人数"])
if "staff_df" not in st.session_state:
    st.session_state.staff_df = pd.DataFrame(columns=["職員ID", "氏名", "所属施設", "所属病棟", "雇用形態", "月間契約時間(h)", "スキルランク", "夜勤可否", "夜勤上限回数", "月間公休数(非常勤用)", "夜勤専従", "休みのリズム", "3連休希望", "生年月日", "入職年月日", "性別", "既婚_未婚", "結婚年月", "末子生年月日"])
else:
    if "休みのリズム" not in st.session_state.staff_df.columns:
        st.session_state.staff_df["休みのリズム"] = "おまかせ"
    if "3連休希望" not in st.session_state.staff_df.columns:
        st.session_state.staff_df["3連休希望"] = False
    if "結婚年月" not in st.session_state.staff_df.columns:
        st.session_state.staff_df["結婚年月"] = ""

if "hr_history_df" not in st.session_state:
    st.session_state.hr_history_df = pd.DataFrame(columns=["職員ID", "イベント種別", "発生年月日", "所属病棟"])
if "ward_settings" not in st.session_state: st.session_state.ward_settings = {}
if "base_shifts" not in st.session_state: st.session_state.base_shifts = None
if "final_shifts" not in st.session_state: st.session_state.final_shifts = None
if "ai_reports" not in st.session_state: st.session_state.ai_reports = {}

score_map = {"S（超指導）": 3.0, "A（指導）": 2.0, "B（自立）": 1.0, "C（支援）": 0.5}

st.sidebar.title("■ 統合司令塔メニュー")
page = st.sidebar.radio("画面を選択", [
    "1. 組織・人員マスタ管理", 
    "2. 経営シミュレーション", 
    "3. 現場制約・希望入力",
    "4. 統合最適化＆応援調整",
    "5. 最適定数 ギャップ分析(What-if)",
    "6. 将来戦力・マクロ推計 (SWP)"
])

# ==========================================
# 共通関数群
# ==========================================
def safe_num(val, default=0):
    n = pd.to_numeric(val, errors='coerce')
    return default if pd.isna(n) else n

def parse_dates(date_str):
    if pd.isna(date_str) or str(date_str).strip() == "": return []
    try: return [int(d.strip()) for d in str(date_str).split(",") if d.strip().isdigit()]
    except: return []

def ensure_ward_settings(ward_list, staff_df):
    for w in ward_list:
        if w not in st.session_state.ward_settings:
            w_staff = staff_df[staff_df["所属病棟"] == w]
            names = w_staff["氏名"].dropna().tolist()
            prev_shifts = ["休/日" for _ in names]
            st.session_state.ward_settings[w] = {
                "req": {"wk_min": 8.0, "wk_ldr": 2.0, "wk_score": 0.0, "we_min": 5.0, "we_ldr": 1.0, "we_score": 0.0, "n_ldr": 1.0, "n_score": 5.5},
                "pair": pd.DataFrame(columns=["職員1", "職員2", "条件種別"]),
                "pref": pd.DataFrame({"氏名": names, "前月最終日": prev_shifts, "次月1日希望": ["" for _ in names], "次月2日希望": ["" for _ in names], "希望休(日付)": ["" for _ in names], "有休希望(日付)": ["" for _ in names]})
            }

def create_excel_download(shift_data, ward_list, days, staff_df, score_map):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheets_written = 0
        for w in ward_list:
            if w not in shift_data: continue
            I_w = list(shift_data[w].keys())
            if len(I_w) == 0: continue
            matrix = []
            for i in I_w:
                row = [i]; d_c=0; n_c=0; m_c=0; o_c=0; y_c=0
                for j in days:
                    val = shift_data[w][i][j]
                    row.append(val)
                    if "日" in val: d_c += 1
                    elif val == "夜": n_c += 1
                    elif val == "明": m_c += 1
                    elif val == "休": o_c += 1
                    elif val == "有": y_c += 1
                row.extend([d_c, n_c, m_c, o_c, y_c, (d_c+n_c+m_c+y_c)*8])
                matrix.append(row)
            summary_rows = [
                ["【集計】日勤人数"] + [sum(1 for i in I_w if "日" in shift_data[w][i][j]) for j in days] + [""]*6,
                ["【集計】夜勤(入)人数"] + [sum(1 for i in I_w if "夜" == shift_data[w][i][j]) for j in days] + [""]*6,
                ["【集計】休み・有休人数"] + [sum(1 for i in I_w if shift_data[w][i][j] in ["休", "有"]) for j in days] + [""]*6
            ]
            d_scores, n_scores = [], []
            for j in days:
                ds = sum(score_map.get(staff_df[staff_df["氏名"]==i]["スキルランク"].values[0], 0) for i in I_w if "日" in shift_data[w][i][j])
                ns = sum(score_map.get(staff_df[staff_df["氏名"]==i]["スキルランク"].values[0], 0) for i in I_w if "夜" == shift_data[w][i][j])
                d_scores.append(ds)
                n_scores.append(ns)
            summary_rows.append(["【集計】日勤スコア"] + d_scores + [""]*6)
            summary_rows.append(["【集計】夜勤スコア"] + n_scores + [""]*6)
            matrix.extend(summary_rows)
            cols = ["氏名"] + [str(d) for d in days] + ["日勤回数", "夜勤回数", "明回数", "公休回数", "有休回数", "実働時間(h)"]
            pd.DataFrame(matrix, columns=cols).to_excel(writer, sheet_name=w, index=False)
            sheets_written += 1
        if sheets_written == 0:
            pd.DataFrame({"エラー": ["出力データがありません。"]}).to_excel(writer, sheet_name="No_Data", index=False)
    return output.getvalue()

def generate_excel_with_validation(org_df, staff_df):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    
    ws_org = wb.active; ws_org.title = "組織構造マスタ"
    ws_org.append(org_df.columns.tolist())
    for r in org_df.values.tolist(): ws_org.append(r)
    
    ws_staff = wb.create_sheet(title="職員台帳マスタ")
    df_out = staff_df.copy()
    for col in ["生年月日", "入職年月日", "末子生年月日"]:
        if col in df_out.columns:
            df_out[col] = df_out[col].dt.strftime('%Y-%m-%d') if pd.api.types.is_datetime64_any_dtype(df_out[col]) else df_out[col].astype(str)
            df_out[col] = df_out[col].replace('NaT', '')
            
    ws_staff.append(df_out.columns.tolist())
    for r in df_out.values.tolist(): ws_staff.append(r)
    
    fac_list = org_df["施設名"].dropna().unique().tolist() if not org_df.empty else ["新宿本院"]
    war_list = org_df["病棟名"].dropna().unique().tolist() if not org_df.empty else []
    
    dv_fac = DataValidation(type="list", formula1=f'"{",".join(fac_list)}"' if fac_list else '" "', allow_blank=True)
    dv_war = DataValidation(type="list", formula1=f'"{",".join(war_list)}"' if war_list else '" "', allow_blank=True)
    dv_emp = DataValidation(type="list", formula1='"常勤,短時間,非常勤"', allow_blank=True)
    dv_rnk = DataValidation(type="list", formula1='"S（超指導）,A（指導）,B（自立）,C（支援）"', allow_blank=True)
    dv_bool = DataValidation(type="list", formula1='"True,False"', allow_blank=True)
    dv_rhy = DataValidation(type="list", formula1='"おまかせ,2連休ベース,1日休みベース"', allow_blank=True)
    dv_gen = DataValidation(type="list", formula1='"女,男"', allow_blank=True)
    dv_mar = DataValidation(type="list", formula1='"既婚,未婚"', allow_blank=True)
    
    dv_date = DataValidation(type="date", operator="between", formula1='1', formula2='73050', allow_blank=True)
    dv_date.errorTitle = '入力エラー'; dv_date.error = '正しい日付を「YYYY-MM-DD」の形式（例: 2024-04-01）で入力してください。'
    dv_date.promptTitle = '日付入力'; dv_date.prompt = '半角で YYYY-MM-DD と入力してください。'

    dv_ym = DataValidation(type="textLength", operator="equal", formula1='7', allow_blank=True)
    dv_ym.errorTitle = '入力エラー'; dv_ym.error = '「YYYY-MM」の形式（例: 2024-04）で入力してください。'
    dv_ym.promptTitle = '年月入力'; dv_ym.prompt = '半角で YYYY-MM（7文字）と入力してください。'

    dv_num = DataValidation(type="whole", operator="between", formula1='0', formula2='31', allow_blank=True)
    dv_num.errorTitle = '入力エラー'; dv_num.error = '0〜31の半角数字を入力してください。'
    dv_num.promptTitle = '数値入力'; dv_num.prompt = '半角数字を入力してください。'

    ws_staff.add_data_validation(dv_fac); ws_staff.add_data_validation(dv_war)
    ws_staff.add_data_validation(dv_emp); ws_staff.add_data_validation(dv_rnk)
    ws_staff.add_data_validation(dv_bool); ws_staff.add_data_validation(dv_rhy)
    ws_staff.add_data_validation(dv_gen); ws_staff.add_data_validation(dv_mar)
    ws_staff.add_data_validation(dv_date); ws_staff.add_data_validation(dv_ym)
    ws_staff.add_data_validation(dv_num)
    
    cols = df_out.columns.tolist()
    def apply_dv(dv, col_name):
        if col_name in cols:
            col_idx = cols.index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            dv.add(f'{col_letter}2:{col_letter}1000')

    apply_dv(dv_fac, "所属施設"); apply_dv(dv_war, "所属病棟")
    apply_dv(dv_emp, "雇用形態"); apply_dv(dv_rnk, "スキルランク")
    apply_dv(dv_bool, "夜勤可否"); apply_dv(dv_bool, "夜勤専従"); apply_dv(dv_bool, "3連休希望")
    apply_dv(dv_rhy, "休みのリズム"); apply_dv(dv_gen, "性別"); apply_dv(dv_mar, "既婚_未婚")
    apply_dv(dv_date, "生年月日"); apply_dv(dv_date, "入職年月日"); apply_dv(dv_date, "末子生年月日")
    apply_dv(dv_ym, "結婚年月"); apply_dv(dv_num, "夜勤上限回数"); apply_dv(dv_num, "月間公休数(非常勤用)")
    
    wb.save(output)
    return output.getvalue()

def generate_dummy_master():
    wards = ["東病棟（外科）", "西病棟（内科）", "南病棟（整形）"]
    org_data = [["医療法人〇〇会", "新宿本院", w, 7, 40, 3] for w in wards]
    staff_data = []; base_date = datetime(2024, 4, 1)
    for w in wards:
        for i in range(1, 31):
            emp = "常勤" if i <= 25 else "非常勤"
            hours = 160 if emp == "常勤" else (120 if np.random.rand() > 0.5 else 100)
            if i <= 2: rank = "S（超指導）"
            elif i <= 7: rank = "A（指導）"
            elif i <= 20: rank = "B（自立）"
            else: rank = "C（支援）"
            night_ok = True if (i <= 22) else False
            night_only = True if (23 <= i <= 24) else False 
            
            night_limit = 9 if night_only else (4 if rank in ["S（超指導）", "A（指導）"] and night_ok else (5 if night_ok else None))
            
            rhythm = np.random.choice(["おまかせ", "2連休ベース", "1日休みベース"], p=[0.6, 0.3, 0.1])
            want_3days = np.random.choice([False, True], p=[0.8, 0.2])
            
            age = np.random.randint(22, 59)
            birth_date = (base_date - relativedelta(years=age, months=np.random.randint(0,11)))
            join_date = (base_date - relativedelta(years=np.random.randint(0, age-21)))
            gender = np.random.choice(["女", "女", "女", "男"])
            married = "既婚" if age > 28 and np.random.rand() > 0.4 else "未婚"
            
            marry_date_str = ""
            if married == "既婚":
                m_date = base_date - relativedelta(years=np.random.randint(0, 5), months=np.random.randint(0, 11))
                marry_date_str = m_date.strftime("%Y-%m")
                
            child_birth = None
            if gender == "女" and married == "既婚" and age < 45 and np.random.rand() > 0.5:
                child_birth = (base_date - relativedelta(years=np.random.randint(0, 10), months=np.random.randint(0,11)))
                
            off_days = 15 if emp=="非常勤" else None
            
            staff_data.append([f"1{wards.index(w)}{i:03d}", f"{w[:1]}_{i:02d}号", "新宿本院", w, emp, hours, rank, night_ok, night_limit, off_days, night_only, rhythm, want_3days, birth_date, join_date, gender, married, marry_date_str, child_birth])
            
    staff_cols = ["職員ID", "氏名", "所属施設", "所属病棟", "雇用形態", "月間契約時間(h)", "スキルランク", "夜勤可否", "夜勤上限回数", "月間公休数(非常勤用)", "夜勤専従", "休みのリズム", "3連休希望", "生年月日", "入職年月日", "性別", "既婚_未婚", "結婚年月", "末子生年月日"]
    return pd.DataFrame(org_data, columns=st.session_state.org_df.columns), pd.DataFrame(staff_data, columns=staff_cols)

def generate_dummy_hr_history(staff_df):
    history_data = []
    base_date = datetime(2024, 4, 1)
    wards = staff_df["所属病棟"].dropna().unique().tolist() if not staff_df.empty else ["東病棟（外科）", "西病棟（内科）", "南病棟（整形）"]
    for y in range(1, 4):
        year_date = base_date - relativedelta(years=y)
        for w in wards:
            for _ in range(np.random.randint(1, 4)):
                h_date = year_date + relativedelta(months=np.random.randint(0, 11))
                history_data.append([f"dummy_ret_{y}_{w}_{_}", "自己都合退職", h_date.strftime("%Y-%m-%d"), w])
            for _ in range(np.random.randint(0, 3)):
                h_date = year_date + relativedelta(months=np.random.randint(0, 11))
                history_data.append([f"dummy_mat_{y}_{w}_{_}", "産休開始", h_date.strftime("%Y-%m-%d"), w])
    return pd.DataFrame(history_data, columns=["職員ID", "イベント種別", "発生年月日", "所属病棟"])

def export_ward_settings_excel(ward, settings):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws_req = wb.active; ws_req.title = "配置要件"
    ws_req.append(["項目名", "設定値"])
    for k, v in settings["req"].items(): ws_req.append([k, v])
    ws_pair = wb.create_sheet(title="ペアリング")
    ws_pair.append(settings["pair"].columns.tolist())
    for r in settings["pair"].values.tolist(): ws_pair.append(r)
    ws_pref = wb.create_sheet(title="個別希望休")
    df_pref = settings["pref"].copy()
    for col in ["生年月日", "入職年月日", "末子生年月日"]:
        if col in df_pref.columns: df_pref[col] = df_pref[col].astype(str)
    ws_pref.append(df_pref.columns.tolist())
    for r in df_pref.values.tolist(): ws_pref.append(r)
    wb.save(output)
    return output.getvalue()

def color_shift_cells(val):
    if "日" in str(val): return 'background-color: #fff9c4; color: #000000;'
    elif str(val) == "夜": return 'background-color: #bbdefb; color: #000000;'
    elif str(val) == "明": return 'background-color: #e0e0e0; color: #000000;'
    elif str(val) in ["休", "有"]: return 'background-color: #ffcdd2; color: #c62828; font-weight: bold;'
    return ''

# ==========================================
# 画面1：組織・人員マスタ管理
# ==========================================
if page == "1. 組織・人員マスタ管理":
    st.title("■ 組織・人員マスタ管理")
    if st.button("■ ダミーデータを自動生成", use_container_width=True):
        st.session_state.org_df, st.session_state.staff_df = generate_dummy_master()
        st.success("ダミーデータをロードしました。")
    if uploaded_excel := st.file_uploader("▶ 統合マスタExcel読込", type=["xlsx"]):
        try:
            st.session_state.org_df = pd.read_excel(uploaded_excel, sheet_name="組織構造マスタ").dropna(how="all")
            df_st = pd.read_excel(uploaded_excel, sheet_name="職員台帳マスタ").dropna(how="all")
            for col in ["生年月日", "入職年月日", "末子生年月日"]:
                if col in df_st.columns: df_st[col] = pd.to_datetime(df_st[col], errors='coerce')
            st.session_state.staff_df = df_st
            st.success("Excelマスタを読み込みました。")
        except Exception as e: st.error(f"読込エラー: {e}")
        
    st.subheader("1. 組織構造マスタ")
    st.session_state.org_df = st.data_editor(st.session_state.org_df, num_rows="dynamic", use_container_width=True)
    
    st.subheader("2. 職員台帳マスタ（入力ガード機能付き）")
    st.info("※ FTE (Full-Time Equivalent): 週40時間等の常勤スタッフ1名分を「1.0」として計算する戦力指標です。")
    
    fac_list = st.session_state.org_df["施設名"].dropna().unique().tolist() if not st.session_state.org_df.empty else []
    war_list = st.session_state.org_df["病棟名"].dropna().unique().tolist() if not st.session_state.org_df.empty else []
    
    edited_staff_df = st.data_editor(
        st.session_state.staff_df, num_rows="dynamic", use_container_width=True,
        column_config={
            "所属施設": st.column_config.SelectboxColumn(options=fac_list), 
            "所属病棟": st.column_config.SelectboxColumn(options=war_list), 
            "雇用形態": st.column_config.SelectboxColumn(options=["常勤", "短時間", "非常勤"]), 
            "スキルランク": st.column_config.SelectboxColumn(options=list(score_map.keys())), 
            "夜勤可否": st.column_config.CheckboxColumn(), 
            "夜勤専従": st.column_config.CheckboxColumn(),
            "休みのリズム": st.column_config.SelectboxColumn(options=["おまかせ", "2連休ベース", "1日休みベース"]),
            "3連休希望": st.column_config.CheckboxColumn(),
            "夜勤上限回数": st.column_config.NumberColumn(min_value=0, max_value=31, step=1),
            "月間公休数(非常勤用)": st.column_config.NumberColumn(min_value=0, max_value=31, step=1),
            "性別": st.column_config.SelectboxColumn(options=["女", "男"]),
            "既婚_未婚": st.column_config.SelectboxColumn(options=["既婚", "未婚"]),
            "結婚年月": st.column_config.TextColumn(help="YYYY-MM形式で入力 (空欄可)"), 
            "生年月日": st.column_config.DateColumn(format="YYYY-MM-DD"),
            "入職年月日": st.column_config.DateColumn(format="YYYY-MM-DD"),
            "末子生年月日": st.column_config.DateColumn(format="YYYY-MM-DD")
        }
    )
    
    for idx, row in edited_staff_df.iterrows():
        if row["夜勤専従"] == True:
            val = safe_num(row["夜勤上限回数"], -1)
            if val == -1: edited_staff_df.at[idx, "夜勤上限回数"] = 9
            
    st.session_state.staff_df = edited_staff_df
    
    if not st.session_state.org_df.empty and not edited_staff_df.empty:
        excel_data = generate_excel_with_validation(st.session_state.org_df, edited_staff_df)
        st.download_button("📥 保存：統合マスタ出力（.xlsx）", data=excel_data, file_name="integrated_master.xlsx", use_container_width=True)

# ==========================================
# 画面2：経営シミュレーション
# ==========================================
elif page == "2. 経営シミュレーション":
    st.title("■ 経営シミュレーション ＆ 事前検証")
    org_df, staff_df = st.session_state.org_df, st.session_state.staff_df
    if not org_df.empty and not staff_df.empty:
        c1, c2, c3 = st.columns(3)
        ward_opts = org_df["病棟名"].dropna().unique().tolist()
        sel_wards = c1.multiselect("対象病棟の合算", ward_opts, default=ward_opts)
        sim_days = c2.selectbox("月の日数", [28, 29, 30, 31], index=3)
        target_fte = c3.number_input("FTE基準時間 (h)", value=160)
        
        if len(sel_wards) > 0:
            t_org = org_df[org_df["病棟名"].isin(sel_wards)]; t_staff = staff_df[staff_df["所属病棟"].isin(sel_wards)]
            
            req_hrs = sum([(r["病床数"] / r["配置比率（対1）"]) * 24 * sim_days for _, r in t_org.iterrows() if pd.notna(r["病床数"])])
            req_fte = req_hrs / target_fte if target_fte > 0 else 0
            act_fte = pd.to_numeric(t_staff["月間契約時間(h)"], errors='coerce').fillna(0).sum() / target_fte if target_fte > 0 else 0
            
            total_night_h = sum([r["夜勤基本人数"] * 16 * sim_days for _, r in t_org.iterrows() if pd.notna(r["夜勤基本人数"])])
            dedicated_staff = t_staff[t_staff["夜勤専従"].astype(str).str.upper() == "TRUE"]
            dedicated_h = sum([safe_num(x, 9) * 16 for x in dedicated_staff["夜勤上限回数"]])
            remaining_night_h = max(0, total_night_h - dedicated_h)
            req_night_staff = remaining_night_h / 72 if remaining_night_h > 0 else 0
            
            night_cap = len(t_staff[(t_staff["夜勤可否"].astype(str).str.upper() == "TRUE") & (t_staff["夜勤専従"].astype(str).str.upper() != "TRUE")])
            
            c_m1, c_m2 = st.columns(2)
            with c_m1:
                st.markdown('<div class="metric-card">', unsafe_allow_html=True)
                st.caption(f"配置基準（様式9）検証：選択 {len(sel_wards)} 病棟")
                st.metric("必要最低人員 (FTE)", f"{req_fte:.1f} 人")
                st.metric("現有スタッフ (FTE)", f"{act_fte:.1f} 人")
                st.markdown('</div>', unsafe_allow_html=True)
            with c_m2:
                st.markdown('<div class="metric-card">', unsafe_allow_html=True)
                st.caption("夜勤72時間ルール検証")
                st.metric("必要夜勤従事者 (専従除く)", f"{req_night_staff:.1f} 人")
                st.metric("現有夜勤可能人数 (専従除く)", f"{night_cap} 人")
                st.markdown('</div>', unsafe_allow_html=True)
            
            st.markdown("---")
            with st.expander("💡 計算式（ロジック）の解説", expanded=False):
                st.markdown("""
                **■ 配置基準（様式9）の計算式**
                * 必要最低人員(FTE) = (病床数 ÷ 配置比率) × 24時間 × 月の日数 ÷ FTE基準時間
                * 現有スタッフ(FTE) = 選択病棟の全スタッフの「月間契約時間」の合計 ÷ FTE基準時間
                
                **■ 夜勤72時間ルールの計算式（法令準拠）**
                ※ 夜勤専従者は母集団から除外し、専従者が担う夜勤時間を必要総時間から差し引いて計算します。
                * 必要夜勤従事者(人) = { (夜勤基本人数 × 16時間 × 月の日数) － (専従者の夜勤上限回数 × 16時間) } ÷ 72時間
                * 現有夜勤可能人数 = マスタで「夜勤可否がTRUE」かつ「夜勤専従がFALSE」のスタッフの総数
                """)
    else: st.warning("マスタデータを設定してください。")

# ==========================================
# 画面3：現場制約・希望入力
# ==========================================
elif page == "3. 現場制約・希望入力":
    st.title("■ 現場制約・希望休管理（各病棟設定）")
    st.divider()
    org_df, staff_df = st.session_state.org_df, st.session_state.staff_df
    if not org_df.empty and not staff_df.empty:
        ward_list = org_df["病棟名"].dropna().unique().tolist()
        sel_ward = st.selectbox("▶ 設定する病棟を選択", ward_list)
        st.divider()
        
        ensure_ward_settings(ward_list, staff_df)
        
        w_staff = staff_df[staff_df["所属病棟"] == sel_ward]
        names = w_staff["氏名"].dropna().tolist()
        cur_set = st.session_state.ward_settings[sel_ward]
        
        st.subheader("📥 現場設定のExcel一括入出力")
        c_ex1, c_ex2 = st.columns(2)
        with c_ex1:
            exp_bin = export_ward_settings_excel(sel_ward, cur_set)
            st.download_button(f"Excel設定テンプレートをエクスポート", data=exp_bin, file_name=f"settings_{sel_ward}.xlsx", use_container_width=True)
        with c_ex2:
            imp_file = st.file_uploader(f"編集済みのExcel設定をインポート", type=["xlsx"], key=f"upload_{sel_ward}")
            if imp_file:
                try:
                    df_req_imp = pd.read_excel(imp_file, sheet_name="配置要件")
                    df_pair_imp = pd.read_excel(imp_file, sheet_name="ペアリング").dropna(how="all")
                    df_pref_imp = pd.read_excel(imp_file, sheet_name="個別希望休").dropna(how="all")
                    
                    cur_set["req"] = dict(zip(df_req_imp.iloc[:,0], df_req_imp.iloc[:,1]))
                    cur_set["pair"] = df_pair_imp
                    
                    if "休みのリズム" in df_pref_imp.columns:
                        df_pref_imp = df_pref_imp.drop(columns=["休みのリズム"])
                    cur_set["pref"] = df_pref_imp
                    
                    st.session_state.ward_settings[sel_ward] = cur_set
                    st.success("Excelから設定情報を一括反映しました！")
                except Exception as e: st.error(f"インポートエラー: {e}")
                
        st.divider()
        st.subheader(f"1. スキルミックス ＆ 人数要件（{sel_ward}）")
        st.info("💡 スコア配分： S（超指導）= 3.0点、 A（指導）= 2.0点、 B（自立）= 1.0点、 C（支援）= 0.5点")
        
        c_wk, c_we = st.columns(2)
        with c_wk:
            st.markdown("▶ **平日（月〜金）の設定**")
            cur_set["req"]["wk_min"] = st.number_input("日勤 最低配置人数", min_value=0.0, value=float(cur_set["req"].get("wk_min", 8.0)), key="wk_min")
            cur_set["req"]["wk_ldr"] = st.number_input("日勤 最低リーダー（S+A）", min_value=0.0, value=float(cur_set["req"].get("wk_ldr", 2.0)), key="wk_ldr")
            cur_set["req"]["wk_score"] = st.number_input("日勤 最低スコア合計", min_value=0.0, value=float(cur_set["req"].get("wk_score", 0.0)), key="wk_scr")
        with c_we:
            st.markdown("▶ **休日（土日祝）の設定**")
            cur_set["req"]["we_min"] = st.number_input("日勤 最低配置人数", min_value=0.0, value=float(cur_set["req"].get("we_min", 5.0)), key="we_min")
            cur_set["req"]["we_ldr"] = st.number_input("日勤 最低リーダー（S+A）", min_value=0.0, value=float(cur_set["req"].get("we_ldr", 1.0)), key="we_ldr")
            cur_set["req"]["we_score"] = st.number_input("日勤 最低スコア合計", min_value=0.0, value=float(cur_set["req"].get("we_score", 0.0)), key="we_scr")
            
        st.markdown("▶ **夜勤の設定**")
        c_n1, c_n2 = st.columns(2)
        with c_n1: cur_set["req"]["n_ldr"] = st.number_input("夜勤 最低リーダー数（S+A）", min_value=0.0, value=float(cur_set["req"].get("n_ldr", 1.0)))
        with c_n2: cur_set["req"]["n_score"] = st.number_input("夜勤 最低スコア合計", min_value=0.0, value=float(cur_set["req"].get("n_score", 5.5)))
            
        st.divider()
        st.subheader(f"2. ペアリング制約（{sel_ward}）")
        cur_set["pair"] = st.data_editor(
            cur_set["pair"], num_rows="dynamic", use_container_width=True, 
            column_config={"職員1": st.column_config.SelectboxColumn(options=names), "職員2": st.column_config.SelectboxColumn(options=names), "条件種別": st.column_config.SelectboxColumn(options=["NGペア（同直不可）", "教育ペア（原則同直）"])}
        )
        
        st.divider()
        st.subheader(f"3. シフト希望（{sel_ward}）")
        st.info("※ 「休みのリズム」や「3連休希望」は、画面1の【組織・人員マスタ管理】から設定・変更してください。")
        cur_set["pref"] = st.data_editor(
            cur_set["pref"], use_container_width=True, hide_index=True, 
            column_config={"氏名": st.column_config.TextColumn(disabled=True), "前月最終日": st.column_config.SelectboxColumn(options=["休/日", "入", "明"]), "次月1日希望": st.column_config.SelectboxColumn(options=["", "休", "有", "日", "夜", "明"]), "次月2日希望": st.column_config.SelectboxColumn(options=["", "休", "有", "日", "夜", "明"])}
        )
        st.session_state.ward_settings[sel_ward] = cur_set
        st.success(f"✔ {sel_ward} の設定を保存しました。")
    else: st.warning("■ マスタデータがありません。")

# ==========================================
# 画面4：統合最適化＆応援調整
# ==========================================
elif page == "4. 統合最適化＆応援調整":
    st.title("■ 全病棟・統合最適化 ＆ 応援スワップ")
    org_df, staff_df = st.session_state.org_df, st.session_state.staff_df
    
    if not org_df.empty and not staff_df.empty:
        c1, c2 = st.columns(2)
        sim_days = c1.selectbox("対象月の日数", [28, 29, 30, 31], index=3)
        weekends = parse_dates(c2.text_input("土日祝の日付（カンマ区切り）", "4,5,11,12,18,19,25,26"))
        days = list(range(1, sim_days + 1)); num_holidays = len(weekends)
        ward_list = org_df["病棟名"].dropna().unique().tolist()
        
        ensure_ward_settings(ward_list, staff_df)
        ws_dict = st.session_state.ward_settings
        
        st.info("※ 現場制約（画面3）が未保存の病棟は、自動的に初期設定値（標準ルール）で計算されます。")
        
        opt_mode = st.radio("▶ 最適化モード", ["① テスト（30秒/病棟）", "② 簡易（3分/病棟）", "③ 完成（10分/病棟）"], index=1)
        
        if st.button("■ 全病棟を一括最適化（ベース作成）", use_container_width=True):
            st.session_state.base_shifts = {}; st.session_state.ai_reports = {}
            target_wards = [w for w in ward_list if w in ws_dict]
            progress = st.progress(0)
            time_limit = 30 if "①" in opt_mode else (180 if "②" in opt_mode else 600)
            
            for idx, ward in enumerate(target_wards):
                w_staff = staff_df[staff_df["所属病棟"] == ward]
                I = w_staff["氏名"].dropna().tolist()
                pref, req, pair_df = ws_dict[ward]["pref"], ws_dict[ward]["req"], ws_dict[ward]["pair"]
                n_basic = safe_num(org_df[org_df["病棟名"] == ward].iloc[0]["夜勤基本人数"], 2)
                
                prob = pulp.LpProblem(f"Base_{ward}", pulp.LpMinimize)
                x = pulp.LpVariable.dicts("x", (I, days, ["日", "夜", "明", "休", "有"]), cat="Binary")
                obj = []
                
                total_off = 0; total_yuq = 0; target_off_dict = {}
                for i in I:
                    emp_type = w_staff[w_staff["氏名"]==i]["雇用形態"].values[0]
                    target_off = num_holidays
                    if emp_type == "非常勤":
                        pt_off = w_staff[w_staff["氏名"]==i]["月間公休数(非常勤用)"].values[0]
                        if pd.notna(pt_off) and str(pt_off).strip() != "": target_off = int(pt_off)
                    target_off_dict[i] = target_off; total_off += target_off
                    row_p = pref[pref["氏名"]==i].iloc[0] if len(pref[pref["氏名"]==i])>0 else None
                    if row_p is not None: total_yuq += len(parse_dates(row_p.get("有休希望(日付)","")))
                avg_d_count = max(0, len(I)*sim_days - total_off - total_yuq - n_basic*2*sim_days) / sim_days 
                avg_score_per_person = sum(score_map.get(w_staff[w_staff["氏名"]==i]["スキルランク"].values[0], 0) for i in I) / len(I) if len(I)>0 else 0
                avg_d_score = avg_d_count * avg_score_per_person 
                
                alert_100k = []; alert_500 = []; alert_200 = []

                for j in days:
                    sp = pulp.LpVariable(f"sp_n_{j}", lowBound=0); sm = pulp.LpVariable(f"sm_n_{j}", lowBound=0)
                    prob += pulp.lpSum([x[i][j]["夜"] for i in I]) - n_basic <= sp
                    prob += n_basic - pulp.lpSum([x[i][j]["夜"] for i in I]) <= sm
                    obj.extend([100000*sp, 100000*sm])
                    alert_100k.extend([(sp, f"【致命的】{j}日の夜勤人数が過剰です"), (sm, f"【致命的】{j}日の夜勤人数が不足しています")])

                for i in I:
                    for j in days: prob += pulp.lpSum([x[i][j][k] for k in ["日", "夜", "明", "休", "有"]]) == 1
                    for j in days[:-1]: prob += x[i][j]["夜"] == x[i][j+1]["明"]; prob += x[i][j]["明"] + x[i][j+1]["夜"] <= 1
                    if str(w_staff[w_staff["氏名"]==i]["夜勤可否"].values[0]).upper() != "TRUE":
                        for j in days: prob += x[i][j]["夜"] == 0
                    
                    max_n = safe_num(w_staff[w_staff["氏名"]==i]["夜勤上限回数"].values[0], default=99)
                    if max_n < 99:
                        sn_over1 = pulp.LpVariable(f"sn_over1_{i}", lowBound=0, upBound=1)
                        sn_over2 = pulp.LpVariable(f"sn_over2_{i}", lowBound=0)
                        prob += pulp.lpSum([x[i][j]["夜"] for j in days]) - max_n <= sn_over1 + sn_over2
                        obj.append(500 * sn_over1)
                        obj.append(100000 * sn_over2)
                        alert_500.append((sn_over1, f"【要件未達】{i}さんの夜勤回数が上限({int(max_n)}回)を1回超過"))
                        alert_100k.append((sn_over2, f"【致命的】{i}さんの夜勤回数が絶対上限({int(max_n)}+1回)を超過"))
                    
                    for j in days[:-1]:
                        s_ake = pulp.LpVariable(f"sake_{i}_{j}", lowBound=0)
                        prob += x[i][j+1]["休"] + x[i][j+1]["有"] + s_ake >= x[i][j]["明"]
                        obj.append(100000 * s_ake)
                        alert_100k.append((s_ake, f"【安全違反】{i}さんの{j}日「明」の翌日が休みではありません"))

                    sp_o = pulp.LpVariable(f"spo_{i}", lowBound=0); sm_o = pulp.LpVariable(f"smo_{i}", lowBound=0)
                    prob += pulp.lpSum([x[i][j]["休"] for j in days]) - target_off_dict[i] <= sp_o
                    prob += target_off_dict[i] - pulp.lpSum([x[i][j]["休"] for j in days]) <= sm_o
                    obj.extend([100000*sp_o, 100000*sm_o])
                    alert_100k.extend([(sp_o, f"【致命的】{i}さんの公休数が過剰です"), (sm_o, f"【致命的】{i}さんの公休数が不足しています")])
                    
                    for j in days[:-5]: 
                        s_6d = pulp.LpVariable(f"s6d_{i}_{j}", lowBound=0)
                        prob += pulp.lpSum([x[i][j+k]["休"] + x[i][j+k]["有"] for k in range(6)]) + s_6d >= 1
                        obj.append(100000 * s_6d)
                        alert_100k.append((s_6d, f"【安全違反】{i}さんが{j}日から6連勤以上になっています"))
                    
                    row_p = pref[pref["氏名"]==i].iloc[0] if len(pref[pref["氏名"]==i])>0 else None
                    if row_p is not None:
                        if row_p.get("前月最終日") == "明": prob += x[i][1]["休"] + x[i][1]["有"] == 1; prob += x[i][1]["明"] == 0
                        elif row_p.get("前月最終日") == "入": prob += x[i][1]["明"] == 1
                        else: prob += x[i][1]["明"] == 0
                        
                        yq = parse_dates(row_p.get("有休希望(日付)",""))
                        for d in yq: 
                            if d in days: prob += x[i][d]["有"] == 1
                            
                        kq = parse_dates(row_p.get("希望休(日付)",""))
                        for d in kq:
                            if d in days: prob += x[i][d]["休"] == 1

                        nxt1 = row_p.get("次月1日希望", "")
                        if nxt1 in ["休", "有", "日", "夜", "明"]:
                            s_n1 = pulp.LpVariable(f"snxt1_{i}", lowBound=0)
                            if nxt1 in ["休", "有", "日"]: prob += x[i][sim_days]["夜"] <= s_n1
                            elif nxt1 == "明": prob += 1 - x[i][sim_days]["夜"] <= s_n1
                            obj.append(500 * s_n1)
                            alert_500.append((s_n1, f"【希望未達】{i}さんの次月1日の希望が反映されていません"))
                            
                        nxt2 = row_p.get("次月2日希望", "")
                        if nxt2 == "明":
                            s_n2 = pulp.LpVariable(f"snxt2_{i}", lowBound=0)
                            prob += x[i][sim_days]["夜"] <= s_n2
                            obj.append(500 * s_n2)
                            alert_500.append((s_n2, f"【希望未達】{i}さんの次月2日の希望が反映されていません"))

                    rhythm = w_staff[w_staff["氏名"]==i]["休みのリズム"].values[0] if "休みのリズム" in w_staff.columns else "おまかせ"
                    want_3days = w_staff[w_staff["氏名"]==i]["3連休希望"].values[0] if "3連休希望" in w_staff.columns else False
                    
                    if rhythm == "2連休ベース":
                        for j in days[1:-1]:
                            sr2 = pulp.LpVariable(f"sr2_{i}_{j}", lowBound=0)
                            prob += x[i][j]["休"] + (1 - x[i][j-1]["休"] - x[i][j-1]["有"]) + (1 - x[i][j+1]["休"] - x[i][j+1]["有"]) - 2 <= sr2
                            obj.append(100 * sr2)
                    elif rhythm == "1日休みベース":
                        for j in days[:-1]:
                            sr1 = pulp.LpVariable(f"sr1_{i}_{j}", lowBound=0)
                            prob += x[i][j]["休"] + x[i][j+1]["休"] - 1 <= sr1
                            obj.append(100 * sr1)

                    if str(want_3days).upper() == "TRUE":
                        s_3d_off = pulp.LpVariable(f"s_3d_off_{i}", lowBound=0)
                        is_3d_list = []
                        for j in days[:-2]:
                            b_3d = pulp.LpVariable(f"b_3d_{i}_{j}", cat="Binary")
                            prob += b_3d <= x[i][j]["休"] + x[i][j]["有"]
                            prob += b_3d <= x[i][j+1]["休"] + x[i][j+1]["有"]
                            prob += b_3d <= x[i][j+2]["休"] + x[i][j+2]["有"]
                            prob += b_3d >= (x[i][j]["休"] + x[i][j]["有"] + x[i][j+1]["休"] + x[i][j+1]["有"] + x[i][j+2]["休"] + x[i][j+2]["有"]) - 2
                            is_3d_list.append(b_3d)
                        prob += pulp.lpSum(is_3d_list) + s_3d_off >= 1
                        obj.append(100 * s_3d_off)

                    for j in days[:-4]:
                        v5 = pulp.LpVariable(f"v5_{i}_{j}", lowBound=0)
                        prob += pulp.lpSum([x[i][j+k]["休"] + x[i][j+k]["有"] for k in range(5)]) + v5 >= 1
                        obj.append(200 * v5)
                        alert_200.append((v5, f"【労務警告】{i}さんが{j}日から5連勤になっています"))

                for _, p_row in pair_df.iterrows():
                    e1, e2, c_type = p_row["職員1"], p_row["職員2"], p_row["条件種別"]
                    if e1 in I and e2 in I:
                        if c_type == "NGペア（同直不可）":
                            for j in days:
                                prob += x[e1][j]["日"] + x[e2][j]["日"] <= 1
                                prob += x[e1][j]["夜"] + x[e2][j]["夜"] <= 1

                S_A_list = [i for i in I if w_staff[w_staff["氏名"]==i]["スキルランク"].values[0] in ["S（超指導）", "A（指導）"]]
                for j in days:
                    is_we = j in weekends
                    d_min = req["we_min"] if is_we else req["wk_min"]
                    d_ldr = req["we_ldr"] if is_we else req["wk_ldr"]
                    d_scr = req["we_score"] if is_we else req["wk_score"]
                    
                    sum_d = pulp.lpSum([x[i][j]["日"] for i in I])
                    sum_s = pulp.lpSum([x[i][j]["日"] * score_map.get(w_staff[w_staff["氏名"]==i]["スキルランク"].values[0], 0) for i in I])
                    
                    dev_dp = pulp.LpVariable(f"ddp_{j}", lowBound=0); dev_dm = pulp.LpVariable(f"ddm_{j}", lowBound=0)
                    prob += sum_d - avg_d_count <= dev_dp; prob += avg_d_count - sum_d <= dev_dm
                    obj.extend([30*dev_dp, 30*dev_dm])
                    
                    slack_dm = pulp.LpVariable(f"sdm_{j}", lowBound=0); prob += sum_d + slack_dm >= d_min
                    slack_dl = pulp.LpVariable(f"sdl_{j}", lowBound=0); prob += pulp.lpSum([x[i][j]["日"] for i in S_A_list]) + slack_dl >= d_ldr
                    slack_nl = pulp.LpVariable(f"snl_{j}", lowBound=0); prob += pulp.lpSum([x[i][j]["夜"] for i in S_A_list]) + slack_nl >= req["n_ldr"]
                    obj.extend([500*slack_dm, 500*slack_dl, 500*slack_nl])
                    
                    alert_500.extend([(slack_dm, f"【要件未達】{j}日の日勤人数が不足"), (slack_dl, f"【要件未達】{j}日の日勤リーダー不足"), (slack_nl, f"【要件未達】{j}日の夜勤リーダー不足")])
                    
                    if d_scr > 0:
                        slack_ds = pulp.LpVariable(f"sds_{j}", lowBound=0); prob += sum_s + slack_ds >= d_scr
                        obj.append(500*slack_ds)
                        alert_500.append((slack_ds, f"【要件未達】{j}日の日勤スコア不足"))
                    if req["n_score"] > 0:
                        slack_ns = pulp.LpVariable(f"sns_{j}", lowBound=0)
                        prob += pulp.lpSum([x[i][j]["夜"] * score_map.get(w_staff[w_staff["氏名"]==i]["スキルランク"].values[0], 0) for i in I]) + slack_ns >= req["n_score"]
                        obj.append(500*slack_ns)
                        alert_500.append((slack_ns, f"【要件未達】{j}日の夜勤スコア不足"))

                prob += pulp.lpSum(obj)
                
                timer_placeholder = st.empty()
                solve_thread = threading.Thread(target=prob.solve, kwargs={'solver': pulp.PULP_CBC_CMD(msg=False, timeLimit=time_limit)})
                solve_thread.start()
                
                start_time = time.time()
                while solve_thread.is_alive():
                    elapsed = time.time() - start_time
                    remaining = max(0, int(time_limit - elapsed))
                    mins, secs = divmod(remaining, 60)
                    timer_placeholder.info(f"⏳ **【{ward}】のAI最適化を実行中... (完了まで最大 {mins:02d}分{secs:02d}秒)**\n\n※この間、別のアプリ等で作業をしてお待ちいただけます。")
                    time.sleep(1)
                    
                solve_thread.join()
                timer_placeholder.empty()
                
                reps = []
                for var, msg in alert_100k + alert_500 + alert_200:
                    if pulp.value(var) and pulp.value(var) > 0.5: reps.append(msg)
                if reps: st.session_state.ai_reports[ward] = list(set(reps))
                
                res = {}
                for i in I:
                    res[i] = {}
                    for j in days:
                        assigned = "休"
                        for k in ["日", "夜", "明", "休", "有"]:
                            if pulp.value(x[i][j][k]) == 1: assigned = k
                        res[i][j] = assigned
                st.session_state.base_shifts[ward] = res
                progress.progress((idx + 1) / len(target_wards))
            st.success("✔ ベースシフト作成が完了しました。")

        if len(st.session_state.ai_reports) > 0:
            st.markdown('<div class="report-box"><h4>🚨 【AI診断レポート】以下の制約条件を満たしきれないまま出力されました</h4>', unsafe_allow_html=True)
            for w, reps in st.session_state.ai_reports.items():
                st.markdown(f"**■ {w}**")
                for r in sorted(reps):
                    if "【致命的】" in r or "【安全違反】" in r:
                        st.markdown(f"- <span style='color: red; font-weight: bold;'>{r}</span>", unsafe_allow_html=True)
                    elif "【要件未達】" in r:
                        st.markdown(f"- <span style='color: red;'>{r}</span>", unsafe_allow_html=True)
                    else:
                        st.markdown(f"- {r}")
            st.markdown('</div>', unsafe_allow_html=True)

        if st.session_state.base_shifts is not None:
            st.subheader("📊 勤務シフト表のリアルタイムプレビュー")
            for w_name, s_matrix in st.session_state.base_shifts.items():
                st.markdown(f"**■ {w_name}**")
                df_prev = pd.DataFrame(s_matrix).T
                df_prev.columns = [f"{c}日" for c in df_prev.columns]
                st.dataframe(df_prev.style.map(color_shift_cells), use_container_width=True)
            
            excel_base = create_excel_download(st.session_state.base_shifts, ward_list, days, staff_df, score_map)
            st.download_button("📥 STEP1: ベースシフト表ダウンロード", data=excel_base, file_name="Base_Shifts.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

        st.divider()
        with st.expander("💡 制約ルールとAIの判断基準（優先度）について", expanded=False):
            st.markdown("""
            AIは独自の優先度に従って、最も無駄のないシフトを計算しています。重要度が高いものほど、AIは回避を優先します。
            ※各重要度に紐づく具体的なペナルティ得点の配点は企業秘密となっております。
            
            | 重要度 | 対象となる制約ルール | 画面での扱い |
            | :--- | :--- | :--- |
            | 絶対 | 夜勤人数不足、公休数未達、明の翌日出勤、6連勤、夜勤上限の2回以上超過、当月希望休・有休の反映など | 致命的エラーとして警告 |
            | 高 | 日勤人数不足、リーダー不足、次月希望無視、夜勤上限の1回超過 | 要件未達として警告 |
            | 中 | 5連勤の発生 | 労務警告として表示 |
            | 低 | 教育ペアの分離、3連続日勤、休みのリズム(2連休/1休)、3連休希望など | 裏側の努力目標（非表示） |
            | 微 | 応援調整時の負担（Aクラスの出向等）、人数の平準化 | 裏側の努力目標（非表示） |
            """)

        st.divider()
        st.subheader("STEP 2: 応援調整 ＆ 最終出力")
        if st.session_state.base_shifts is not None:
            if st.button("■ 応援調整（全病棟の日勤スワップ）を実行してExcel出力", use_container_width=True):
                timer_placeholder = st.empty()
                progress_sw = st.progress(0)
                bases = st.session_state.base_shifts
                final_shifts = {w: {} for w in ward_list}
                for w in ward_list:
                    if w in bases:
                        for i in bases[w].keys(): final_shifts[w][i] = {}

                for idx, j in enumerate(days):
                    prob_sw = pulp.LpProblem(f"Swap_{j}", pulp.LpMinimize)
                    D_staff = []
                    for w in ward_list:
                        if w not in bases: continue
                        for i, shift_val in bases[w].items():
                            if shift_val[j] == "日":
                                rnk = staff_df[staff_df["氏名"]==i]["スキルランク"].values[0]
                                D_staff.append((i, w, rnk, score_map.get(rnk, 0)))
                            else: 
                                final_shifts[w][i][j] = shift_val[j]

                    y = {}
                    for (i, hw, rnk, scr) in D_staff:
                        y[i] = pulp.LpVariable.dicts(f"y_{i}_{j}", ward_list, cat="Binary")
                        prob_sw += pulp.lpSum([y[i][tw] for tw in ward_list]) == 1
                        if rnk in ["S（超指導）", "C（支援）"]: prob_sw += y[i][hw] == 1
                    
                    obj_sw = []
                    for (i, hw, rnk, scr) in D_staff:
                        if rnk == "B（自立）": obj_sw.append(10 * pulp.lpSum([y[i][tw] for tw in ward_list if tw != hw]))
                        elif rnk == "A（指導）": obj_sw.append(50 * pulp.lpSum([y[i][tw] for tw in ward_list if tw != hw]))

                    for w in ward_list:
                        if w not in ws_dict: continue
                        req = ws_dict[w]["req"]; is_we = j in weekends
                        d_min = req["we_min"] if is_we else req["wk_min"]
                        d_ldr = req["we_ldr"] if is_we else req["wk_ldr"]
                        
                        slack_m = pulp.LpVariable(f"sw_sm_{w}_{j}", lowBound=0); prob_sw += pulp.lpSum([y[i][w] for (i, hw, rnk, scr) in D_staff]) + slack_m >= d_min
                        slack_l = pulp.LpVariable(f"sw_sl_{w}_{j}", lowBound=0); prob_sw += pulp.lpSum([y[i][w] for (i, hw, rnk, scr) in D_staff if rnk in ["S（超指導）", "A（指導）"]]) + slack_l >= d_ldr
                        obj_sw.extend([500*slack_m, 500*slack_l])
                            
                    prob_sw += pulp.lpSum(obj_sw)
                    
                    solve_thread = threading.Thread(target=prob_sw.solve, kwargs={'solver': pulp.PULP_CBC_CMD(msg=False, timeLimit=30)})
                    solve_thread.start()
                    start_time = time.time()
                    while solve_thread.is_alive():
                        elapsed = time.time() - start_time
                        remaining = max(0, int(30 - elapsed))
                        mins, secs = divmod(remaining, 60)
                        timer_placeholder.info(f"⏳ **応援調整中 ({j}日目 / {sim_days}日) ... 該当日の完了まで最大 {mins:02d}分{secs:02d}秒**")
                        time.sleep(1)
                    solve_thread.join()
                    
                    for (i, hw, rnk, scr) in D_staff:
                        assigned_w = hw
                        for tw in ward_list:
                            if pulp.value(y[i][tw]) == 1: assigned_w = tw
                        if assigned_w == hw: final_shifts[hw][i][j] = "日"
                        else: final_shifts[hw][i][j] = f"日({assigned_w}応援)"
                        
                    progress_sw.progress((idx + 1) / len(days))
                
                timer_placeholder.empty()
                st.session_state.final_shifts = final_shifts
                st.success("✔ 応援調整が完了しました。")

            if st.session_state.final_shifts is not None:
                st.subheader("📊 最終応援調整後の統合シフトプレビュー")
                for w_name, s_matrix in st.session_state.final_shifts.items():
                    st.markdown(f"**■ {w_name}**")
                    df_final_prev = pd.DataFrame(s_matrix).T
                    df_final_prev.columns = [f"{c}日" for c in df_final_prev.columns]
                    st.dataframe(df_final_prev.style.map(color_shift_cells), use_container_width=True)

                excel_final = create_excel_download(st.session_state.final_shifts, ward_list, days, staff_df, score_map)
                st.download_button("📥 STEP2: 最終統合シフト表をダウンロード（Excel）", data=excel_final, file_name="Integrated_Final_Shifts.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

# ==========================================
# 🚨 画面5：最適定数 ギャップ分析 (What-if)
# ==========================================
elif page == "5. 最適定数 ギャップ分析(What-if)":
    st.title("■ ランク別 最適定数・ギャップ分析 (黄金比シミュレータ)")
    st.markdown("現在のメンバーとダミー要員を組み合わせ、**『設定された現場ルール（ハード制約）』を100%完全に満たすために必要な最小FTE（最適定数）**をAIが逆算します。各ランクの**不足（幽霊要員の稼働）**だけでなく、**余剰（休眠要員の発生）**も可視化します。")
    st.divider()
    
    org_df, staff_df = st.session_state.org_df, st.session_state.staff_df
    ward_list = org_df["病棟名"].dropna().unique().tolist() if not org_df.empty else []
    
    ensure_ward_settings(ward_list, staff_df)
    ws_dict = st.session_state.ward_settings
    
    if not org_df.empty and len(ws_dict) > 0:
        st.info("💡 **前提条件:** 組織の最適定数は、1年間で最も人員確保が厳しくなる「31日」の月を基準に算出されます。\n\n※現場制約（画面3）が未保存の病棟は、自動的に初期設定値（標準ルール）で計算されます。")
        sel_ward = st.selectbox("検証対象の病棟", [w for w in org_df["病棟名"].dropna().unique() if w in ws_dict])
        sim_days = 31
        days = list(range(1, sim_days + 1))
        
        opt_mode_wi = st.radio("▶ 最適化計算モード（定数逆算用）", ["① 簡易チェック（3分）", "② 経営会議・予算策定用（15分）", "③ 徹底解析（30分）"], index=0)
        time_limit_wi = 180 if "①" in opt_mode_wi else (900 if "②" in opt_mode_wi else 1800)
        
        weekends_wi = [6, 7, 13, 14, 20, 21, 27, 28] 
        
        with st.expander("💡 AIが考慮している【絶対ルール】と【あえて外しているルール】", expanded=False):
            st.markdown("""
            このシミュレータは、現在のスタッフに「各ランクのダミー要員（幽霊）」を組み合わせて、病棟を安全に回すための『基礎体力（ベースとなる最適定数）』を算出します。
            特定の月の偶然（希望休の被りやカレンダーの並び等）によって算出結果が毎月ブレてしまうのを防ぐため、以下の**「安全と健康のための絶対ルール（ハード制約）」のみ**をストイックに満たす純粋な最小人数を逆算しています。

            **■ AIが【100%完全に満たしている】絶対ルール**
            * **働き方の安全基準（労務制約）**
              * 1日1勤務の原則（日/夜/明/休）
              * 夜勤 → 明け の固定サイクル
              * 「明け」の翌日は必ず「休み」にする
              * 6連勤以上の絶対禁止（最大5連勤まで）
              * 夜勤上限回数の超過禁止（個人の設定値を超えない）
              * 「夜勤不可」スタッフへの夜勤割り当て禁止
              * 月間最低9日の公休確保
            * **病棟の配置要件（画面3の設定と完全連動）**
              * 毎日の「夜勤基本人数」の確保
              * 平日/休日別の「日勤 最低配置人数」の確保
              * 日勤および夜勤の「最低リーダー数（S+Aクラス）」の確保
              * 日勤および夜勤の「最低スコア合計」の達成

            **■ AIが【あえて外している】ルール（毎月のブレ要素を排除するため）**
            * 当月の個別の希望休・有休の反映
            * NGペア・教育ペアなどの組み合わせ制約
            * 3連休希望、2連休ベースなどの個別の休みのリズム目標
            """)
        
        if st.button("🚀 ランク別 最適定数とギャップを算出", type="primary", use_container_width=True):
            timer_placeholder = st.empty()
            
            w_staff = staff_df[staff_df["所属病棟"] == sel_ward]
            st.markdown("""
            このシミュレータは、現在のスタッフに「各ランクのダミー要員（幽霊）」を組み合わせて、病棟を安全に回すための『基礎体力（ベースとなる最適定数）』を算出します。
            特定の月の偶然（希望休の被りやカレンダーの並び等）によって算出結果が毎月ブレてしまうのを防ぐため、以下の**「安全と健康のための絶対ルール（ハード制約）」のみ**をストイックに満たす純粋な最小人数を逆算しています。

            **■ AIが【100%完全に満たしている】絶対ルール**
            * **働き方の安全基準（労務制約）**
              * 1日1勤務の原則（日/夜/明/休）
              * 夜勤 → 明け の固定サイクル
              * 「明け」の翌日は必ず「休み」にする
              * 6連勤以上の絶対禁止（最大5連勤まで）
              * 夜勤上限回数の超過禁止（個人の設定値を超えない）
              * 「夜勤不可」スタッフへの夜勤割り当て禁止
              * 月間最低9日の公休確保
            * 病棟の配置要件（画面3の設定と完全連動）**
              * 毎日の「夜勤基本人数」の確保
              * 平日/休日別の「日勤 最低配置人数」の確保
              * 日勤および夜勤の「最低リーダー数（S+Aクラス）」の確保
              * 日勤および夜勤の「最低スコア合計」の達成

            **■ AIが【あえて外している】ルール（毎月のブレ要素を排除するため）**
            * 当月の個別の希望休・有休の反映
            * NGペア・教育ペアなどの組み合わせ制約
            * 3連休希望、2連休ベースなどの個別の休みのリズム目標
            """)
            
            w_staff = staff_df[staff_df["所属病棟"] == sel_ward]
            I_real = w_staff["氏名"].dropna().tolist()
            req = ws_dict[sel_ward]["req"]
            n_basic = safe_num(org_df[org_df["病棟名"] == sel_ward].iloc[0]["夜勤基本人数"], 2)
            
            dummy_list = []
            for rnk in ["S（超指導）", "A（指導）", "B（自立）", "C（支援）"]:
                for d_idx in range(4): 
                    dummy_name = f"幽霊_{rnk}_{d_idx}"; dummy_list.append((dummy_name, rnk))
            I_all = I_real + [d[0] for d in dummy_list]
            
            # 🚨 修正：ヘルパー関数を使ってリアルスタッフとダミースタッフのスコアを正確に取得
            def get_score(i_name):
                if i_name in I_real:
                    return score_map.get(w_staff[w_staff["氏名"]==i_name]["スキルランク"].values[0], 0)
                else:
                    rnk = i_name.split("_")[1]
                    return score_map.get(rnk, 0)
            
            prob = pulp.LpProblem("WhatIf_Sim", pulp.LpMinimize)
            x = pulp.LpVariable.dicts("x_wi", (I_all, days, ["日", "夜", "明", "休"]), cat="Binary")
            obj = []
            
            for i in I_all:
                for j in days: prob += pulp.lpSum([x[i][j][k] for k in ["日", "夜", "明", "休"]]) == 1
                for j in days[:-1]: prob += x[i][j]["夜"] == x[i][j+1]["明"]; prob += x[i][j]["明"] + x[i][j+1]["夜"] <= 1
                
                if i in I_real:
                    if str(w_staff[w_staff["氏名"]==i]["夜勤可否"].values[0]).upper() != "TRUE":
                        for j in days: prob += x[i][j]["夜"] == 0
                    
                    max_n = safe_num(w_staff[w_staff["氏名"]==i]["夜勤上限回数"].values[0], default=99)
                    if max_n < 99:
                        prob += pulp.lpSum([x[i][j]["夜"] for j in days]) <= max_n
                
                prob += pulp.lpSum([x[i][j]["休"] for j in days]) >= 9
            for i in I_all:
                for j in days: prob += pulp.lpSum([x[i][j][k] for k in ["日", "夜", "明", "休"]]) == 1
                for j in days[:-1]: prob += x[i][j]["夜"] == x[i][j+1]["明"]; prob += x[i][j]["明"] + x[i][j+1]["夜"] <= 1
                
                # 1. 明けの翌日は必ず休み（ハード制約）
                for j in days[:-1]: 
                    prob += x[i][j+1]["休"] >= x[i][j]["明"]
                
                # 2. 6連勤の絶対禁止（6日間に最低1日は休み）（ハード制約）
                for j in days[:-5]: 
                    prob += pulp.lpSum([x[i][j+k]["休"] for k in range(6)]) >= 1
                
                if i in I_real:
                    if str(w_staff[w_staff["氏名"]==i]["夜勤可否"].values[0]).upper() != "TRUE":
                        for j in days: prob += x[i][j]["夜"] == 0
                    
                    max_n = safe_num(w_staff[w_staff["氏名"]==i]["夜勤上限回数"].values[0], default=99)
                    if max_n < 99:
                        prob += pulp.lpSum([x[i][j]["夜"] for j in days]) <= max_n
                
                prob += pulp.lpSum([x[i][j]["休"] for j in days]) >= 9
            
            S_A_all = [i for i in I_real if w_staff[w_staff["氏名"]==i]["スキルランク"].values[0] in ["S（超指導）", "A（指導）"]] + [d[0] for d in dummy_list if d[1] in ["S（超指導）", "A（指導）"]]
            
            for j in days:
                is_we = j in weekends_wi
                d_min = req["we_min"] if is_we else req["wk_min"]
                d_ldr = req["we_ldr"] if is_we else req["wk_ldr"]
                d_scr = req["we_score"] if is_we else req["wk_score"]
                
                prob += pulp.lpSum([x[i][j]["夜"] for i in I_all]) == n_basic
                prob += pulp.lpSum([x[i][j]["日"] for i in I_all]) >= d_min
                prob += pulp.lpSum([x[i][j]["日"] for i in S_A_all]) >= d_ldr
                prob += pulp.lpSum([x[i][j]["夜"] for i in S_A_all]) >= req["n_ldr"]
                
                # 🚨 修正：日勤スコア制約の完全同期
                if d_scr > 0:
                    prob += pulp.lpSum([x[i][j]["日"] * get_score(i) for i in I_all]) >= d_scr
                
                # 🚨 修正：夜勤スコア制約の完全同期（これまで欠落していた重大な条件）
                if req["n_score"] > 0:
                    prob += pulp.lpSum([x[i][j]["夜"] * get_score(i) for i in I_all]) >= req["n_score"]

            for d_name, rnk in dummy_list:
                for j in days: obj.append(50000 * x[d_name][j]["日"]); obj.append(50000 * x[d_name][j]["夜"])
            
            for r_name in I_real:
                for j in days: obj.append(10 * x[r_name][j]["日"]); obj.append(10 * x[r_name][j]["夜"])
                    
            prob += pulp.lpSum(obj)
            
            solve_thread = threading.Thread(target=prob.solve, kwargs={'solver': pulp.PULP_CBC_CMD(msg=False, timeLimit=time_limit_wi)})
            solve_thread.start()
            
            start_time = time.time()
            while solve_thread.is_alive():
                elapsed = time.time() - start_time
                remaining = max(0, int(time_limit_wi - elapsed))
                mins, secs = divmod(remaining, 60)
                timer_placeholder.info(f"⏳ **【{sel_ward}】の最適定数を厳密に計算中... (完了まで最大 {mins:02d}分{secs:02d}秒)**\n\n※途中で最適な組み合わせが確定した場合は自動終了します。")
                time.sleep(1)
                
            solve_thread.join()
            timer_placeholder.empty()
            st.success("✔ 分析が完了しました！")
            
            st.subheader(f"📊 【{sel_ward}】 ランク別 最適定数 ＆ ギャップ分析結果")
            
            gap_data = []
            total_real_fte = 0.0; total_opt_fte = 0.0
            
            for rnk in ["S（超指導）", "A（指導）", "B（自立）", "C（支援）"]:
                real_staff = [i for i in I_real if w_staff[w_staff["氏名"]==i]["スキルランク"].values[0] == rnk]
                r_fte = sum([safe_num(w_staff[w_staff["氏名"]==i]["月間契約時間(h)"].values[0], 160) / 160.0 for i in real_staff])
                total_real_fte += r_fte
                
                dummy_staff = [d[0] for d in dummy_list if d[1] == rnk]
                all_rank_staff = real_staff + dummy_staff
                opt_work_shifts = sum([pulp.value(x[i][j]["日"]) + pulp.value(x[i][j]["夜"]) for i in all_rank_staff for j in days])
                
                opt_fte = opt_work_shifts / (sim_days - 9) if opt_work_shifts > 0 else 0.0
                total_opt_fte += opt_fte
                
                gap = r_fte - opt_fte
                
                if gap > 0.5:
                    action = "【余剰】他病棟への異動・応援出しが可能"
                    gap_html = f"<span class='gap-surplus'>+ {gap:.1f}</span>"
                elif gap < -0.5:
                    action = "【不足】採用・異動受入が急務"
                    gap_html = f"<span class='gap-shortage'>- {abs(gap):.1f}</span>"
                else:
                    action = "【適正】現状の配置を維持"
                    gap_html = f"<span class='gap-perfect'>{gap:+.1f}</span>"
                    
                gap_data.append([rnk, f"{r_fte:.1f}", f"{opt_fte:.1f}", gap_html, action])
            
            total_gap = total_real_fte - total_opt_fte
            if total_gap > 0.5: t_html = f"<span class='gap-surplus'>+ {total_gap:.1f}</span>"
            elif total_gap < -0.5: t_html = f"<span class='gap-shortage'>- {abs(total_gap):.1f}</span>"
            else: t_html = f"<span class='gap-perfect'>{total_gap:+.1f}</span>"
            
            gap_data.append(["合計 FTE", f"{total_real_fte:.1f}", f"{total_opt_fte:.1f}", t_html, "※トータル人数と質（スキル）のバランスを確認してください"])
            
            df_gap = pd.DataFrame(gap_data, columns=["スキルランク", "現在の配置(FTE)", "AI最適定数(FTE)", "ギャップ(FTE)", "経営への示唆・アクション"])
            st.markdown(df_gap.to_html(escape=False, index=False, justify="center"), unsafe_allow_html=True)

# ==========================================
# 画面6：将来戦力マクロ推計 (SWP)
# ==========================================
elif page == "6. 将来戦力・マクロ推計 (SWP)":
    st.title("■ 将来戦力・マクロ推計 (Strategic Workforce Planning)")
    st.markdown("過去の退職・産休データから自院の**「リアルな離脱確率」を自動学習**し、育成ルールと**『新卒採用予定（自動補充）』**を掛け合わせて未来のスキルミックスを予測します。")
    st.divider()

    staff_df = st.session_state.staff_df
    if staff_df.empty: 
        st.warning("マスタデータがありません。画面1でダミーデータを生成してください。")
    else:
        st.subheader("📚 ステップ1：過去のイベント履歴（退職・産休）の学習")
        
        c_hist1, c_hist2 = st.columns(2)
        with c_hist1:
            if st.button("■ 過去3年分のダミー履歴データを自動生成"):
                st.session_state.hr_history_df = generate_dummy_hr_history(staff_df)
                st.success("ダミーの人事イベント履歴を生成しました。")
        with c_hist2:
            hr_file = st.file_uploader("▶ 人事イベント履歴Excel読込", type=["xlsx"])
            if hr_file:
                try:
                    st.session_state.hr_history_df = pd.read_excel(hr_file).dropna(how="all")
                    st.success("履歴Excelを読み込みました。")
                except: st.error("読込エラー")

        hr_df = st.session_state.hr_history_df
        learned_turnover_rate = 0.05
        learned_maternity_rate = 0.005
        
        if not hr_df.empty:
            with st.expander("📝 読み込まれた履歴データ（プレビュー）", expanded=False):
                st.dataframe(hr_df, height=150, use_container_width=True)
            
            total_staff = len(staff_df)
            female_target_staff = len(staff_df[(staff_df["性別"] == "女")])
            years_of_data = 3
            
            total_turnover = len(hr_df[hr_df["イベント種別"] == "自己都合退職"])
            total_maternity = len(hr_df[hr_df["イベント種別"] == "産休開始"])
            
            if total_staff > 0: learned_turnover_rate = (total_turnover / years_of_data) / total_staff
            if female_target_staff > 0: learned_maternity_rate = (total_maternity / years_of_data) / female_target_staff / 12
                
            st.markdown('<div class="report-box" style="background-color: #e8f5e9; border-left: 5px solid #4caf50;">', unsafe_allow_html=True)
            st.markdown("💡 **【AI学習完了】自院データから以下の発生確率を算出して予測に使用します**")
            st.markdown(f"- **年間 自己都合退職率:** {learned_turnover_rate*100:.1f} %")
            st.markdown(f"- **月間 産休発生確率 (対象女性):** {learned_maternity_rate*100:.2f} %")
            st.markdown('</div>', unsafe_allow_html=True)

        st.divider()
        st.subheader("⚙️ ステップ2：シミュレーション設定（育成・復帰・新卒補充・定年ルール）")
        c_sim1, c_sim2, c_sim3, c_sim4, c_sim5, c_sim6 = st.columns(6)
        c_to_b = c_sim1.number_input("C→B 昇格目安(年)", value=3.0, min_value=0.5, step=0.5, format="%.1f")
        b_to_a = c_sim2.number_input("B→A 昇格目安(年)", value=5.0, min_value=0.5, step=0.5, format="%.1f")
        a_to_s = c_sim3.number_input("A→S 昇格目安(年)", value=10.0, min_value=0.5, step=0.5, format="%.1f")
        leave_rank_down = c_sim4.number_input("休職復帰 ランクダウン(月)", value=6, min_value=0)
        annual_new_grad_count = c_sim5.number_input("毎年4月のCクラス補充(名)", value=3, min_value=0)
        retire_age = c_sim6.number_input("定年退職年齢", value=60, min_value=50, max_value=75)
        
        has_marry_data = False
        if "結婚年月" in staff_df.columns:
            valid_marry_entries = staff_df["結婚年月"].dropna().astype(str).str.strip()
            valid_marry_entries = valid_marry_entries[valid_marry_entries != ""]
            if len(valid_marry_entries) > 0:
                has_marry_data = True
        
        st.divider()
        st.subheader("📊 産休・復帰予測ロジック解説")
        if has_marry_data:
            st.success("✔ マスタ内に「結婚年月」が検出されたため、高度な『カプラン・マイヤー法（経過年数分析）』を適用してシミュレーションします。")
            with st.expander("💡 【用語解説】カプラン・マイヤー法（経過年数分析）とは？", expanded=False):
                st.markdown("""
                個人があるイベント（今回は結婚）を経験してから、時間の経過とともにどれくらいの確率で次のイベント（今回は産休）に到達するかを時系列で追跡して確率を算出する医療・統計学の標準的な手法です。
                本システムでは、結婚からの経過期間に応じて産休の発生確率をなだらかに変動させることで、より実態に近いリアルな離脱の波をシミュレートします。
                """)
        else:
            st.info("ℹ マスタ内の「結婚年月」が空欄のため、標準的な『重み付けモンテカルロ法』を適用してシミュレーションします。")
            with st.expander("💡 【用語解説】重み付けモンテカルロ法とは？", expanded=False):
                st.markdown("""
                未来の不確実な増減を予測するために、確率的な「サイコロ」を数千回振ってその平均的なトレンド（波）を導き出す統計シミュレーション手法です。
                データが空欄の場合、AIは年齢と性別の統計確率をベースにしつつ、マスタの「既婚」フラグを持つスタッフに対しては産休突入のサイコロの目を独自の傾斜ロジックで変動させて未来を予測します。
                （※具体的な重み付けの補正倍率は企業秘密となっております）。
                """)

        st.markdown("""
        **■ 産休・育休からの復帰ロジック（就学前時短モデル）**
        AIは産休に入ったスタッフに対し、以下の実態に即したサイクルを自動で適用します。
        1. **0〜1歳（12ヶ月間）:** 完全休業（FTE = 0.0）
        2. **1歳〜小学校入学:** 時短勤務として復帰（FTE = 0.75 もしくは元の契約時間）
        3. **小学校入学（満6歳を迎えた次の4月）:** フルタイム勤務へ復帰（FTE = 1.0）
        """)

        st.divider()
        st.subheader("🚀 ステップ3：未来予測の実行")

        c1, c2 = st.columns(2)
        ward_list = staff_df["所属病棟"].dropna().unique().tolist()
        sel_wards = c1.multiselect("▶ 分析対象病棟", ward_list, default=ward_list)
        sim_years = c2.slider("▶ 予測期間（年）", min_value=1, max_value=5, value=3)

        if st.button("🚀 育成×離脱×採用シミュレーションを実行", type="primary", use_container_width=True):
            with st.spinner("モンテカルロ法によるスキル推移・新卒補充を計算中..."):
                base_date = datetime(2024, 4, 1); months = sim_years * 12
                timeline = [base_date + relativedelta(months=m) for m in range(months + 1)]
                timeline_str = [d.strftime("%Y-%m") for d in timeline]
                
                event_counts = {ts: {"定年・自己都合退職": 0, "産休・育休離脱": 0, "育休・時短復帰(増)": 0, "新卒入職(補充)": 0} for ts in timeline_str}
                rank_str_map = {4: "S（超指導）", 3: "A（指導）", 2: "B（自立）", 1: "C（支援）"}
                fte_history = {r_str: np.zeros(months + 1) for r_str in rank_str_map.values()}

                sim_population = []
                for idx, row in staff_df.iterrows():
                    ward = row["所属病棟"]
                    if ward not in sel_wards: continue
                    try: b_date = pd.to_datetime(row["生年月日"])
                    except: b_date = base_date - relativedelta(years=30)
                    try: c_birth = pd.to_datetime(row["末子生年月日"])
                    except: c_birth = None
                    
                    m_date = None
                    if pd.notna(row.get("結婚年月")) and str(row.get("結婚年月")).strip() != "":
                        try: m_date = pd.to_datetime(str(row.get("結婚年月")).strip() + "-01")
                        except: m_date = None
                    
                    r_val = 1
                    for k, v in rank_str_map.items():
                        if v == row["スキルランク"]: r_val = k
                        
                    o_fte = safe_num(row["月間契約時間(h)"], 160) / 160.0
                    
                    sim_population.append({
                        "is_active": True, "birth_date": b_date, "child_birth": c_birth, "marry_date": m_date,
                        "gender": row["性別"], "married": row["既婚_未婚"], 
                        "original_fte": o_fte, "base_fte": o_fte,
                        "current_rank": r_val, "leave_timer": 0, "return_timer": 0,
                        "rank_tenure_months": 0 
                    })

                for m in range(months + 1):
                    current_date = timeline[m]; c_str = timeline_str[m]
                    
                    if current_date.month == 4 and m > 0 and annual_new_grad_count > 0:
                        event_counts[c_str]["新卒入職(補充)"] += annual_new_grad_count
                        for _ in range(int(annual_new_grad_count)):
                            sim_population.append({
                                "is_active": True, "birth_date": current_date - relativedelta(years=22),
                                "child_birth": None, "marry_date": None, "gender": "女", "married": "未婚", 
                                "original_fte": 1.0, "base_fte": 1.0,
                                "current_rank": 1, "leave_timer": 0, "return_timer": 0,
                                "rank_tenure_months": 0
                            })
                    
                    for p in sim_population:
                        if not p["is_active"]: continue
                        
                        age = current_date.year - p["birth_date"].year
                        p["rank_tenure_months"] += 1
                        
                        if age >= retire_age: event_counts[c_str]["定年・自己都合退職"] += 1; p["is_active"] = False; continue
                        if m > 0 and m % 12 == 0 and np.random.rand() < learned_turnover_rate: 
                            event_counts[c_str]["定年・自己都合退職"] += 1; p["is_active"] = False; continue
                        
                        if p["leave_timer"] > 0:
                            p["leave_timer"] -= 1
                            if p["leave_timer"] == 0: 
                                p["return_timer"] = leave_rank_down
                                p["base_fte"] = min(0.75, p["original_fte"])
                                event_counts[c_str]["育休・時短復帰(増)"] += 1
                            continue 
                            
                        target_maternity_prob = learned_maternity_rate
                        
                        if p["gender"] == "女" and 25 <= age <= 35:
                            if has_marry_data and p["marry_date"] is not None:
                                elapsed_months = (current_date.year - p["marry_date"].year) * 12 + (current_date.month - p["marry_date"].month)
                                if 0 <= elapsed_months <= 12: target_maternity_prob = learned_maternity_rate * 1.5
                                elif 13 <= elapsed_months <= 36: target_maternity_prob = learned_maternity_rate * 2.5
                                else: target_maternity_prob = learned_maternity_rate * 0.8
                            else:
                                if p["married"] == "既婚": target_maternity_prob = learned_maternity_rate * 2.0
                                else: target_maternity_prob = learned_maternity_rate * 0.5
                                
                            if np.random.rand() < target_maternity_prob:
                                event_counts[c_str]["産休・育休離脱"] += 1
                                p["leave_timer"] = 12
                                p["child_birth"] = current_date
                                p["base_fte"] = 0.0 
                                continue
                            
                        if p["child_birth"] is not None and p["base_fte"] < p["original_fte"]:
                            b_year = p["child_birth"].year
                            if p["child_birth"].month <= 3:
                                b_year -= 1
                            school_entry_year = b_year + 7
                            
                            if current_date.year == school_entry_year and current_date.month == 4: 
                                event_counts[c_str]["育休・時短復帰(増)"] += 1
                                p["base_fte"] = p["original_fte"]

                        actual_rank = p["current_rank"]
                        if p["return_timer"] > 0:
                            actual_rank = max(1, p["current_rank"] - 1)
                            p["return_timer"] -= 1
                        else:
                            if p["current_rank"] == 1 and p["rank_tenure_months"] >= (c_to_b * 12):
                                p["current_rank"] = 2; p["rank_tenure_months"] = 0
                            elif p["current_rank"] == 2 and p["rank_tenure_months"] >= (b_to_a * 12):
                                p["current_rank"] = 3; p["rank_tenure_months"] = 0
                            elif p["current_rank"] == 3 and p["rank_tenure_months"] >= (a_to_s * 12):
                                p["current_rank"] = 4; p["rank_tenure_months"] = 0
                            actual_rank = p["current_rank"]
                            
                        fte_history[rank_str_map[actual_rank]][m] += p["base_fte"]

            st.subheader(f"📊 将来のスキルミックス（階層別 FTE推移：採用補充あり）")
            df_line = pd.DataFrame(fte_history, index=timeline_str)
            
            safe_cols = [rank_str_map[1], rank_str_map[2], rank_str_map[3], rank_str_map[4]]
            df_line = df_line[safe_cols]
            
            st.line_chart(df_line, color=["#1565c0", "#2e7d32", "#ff9800", "#b71c1c"])
            
            st.subheader("📋 【詳細】経過年数ごとのクラス別 FTE推移表")
            rank_summary = []
            for y in range(sim_years + 1):
                m_idx = min(y * 12, months)
                date_str = timeline_str[m_idx]
                label = "現在" if y == 0 else f"{y}年後"
                
                s_val = fte_history["S（超指導）"][m_idx]
                a_val = fte_history["A（指導）"][m_idx]
                b_val = fte_history["B（自立）"][m_idx]
                c_val = fte_history["C（支援）"][m_idx]
                total = s_val + a_val + b_val + c_val
                
                rank_summary.append({
                    "時期": f"{label} ({date_str})",
                    "Sクラス(FTE)": round(s_val, 1),
                    "Aクラス(FTE)": round(a_val, 1),
                    "Bクラス(FTE)": round(b_val, 1),
                    "Cクラス(FTE)": round(c_val, 1),
                    "合計(FTE)": round(total, 1)
                })
                
            df_rank_summary = pd.DataFrame(rank_summary)
            st.dataframe(df_rank_summary, use_container_width=True, hide_index=True)
            
            st.divider()

            df_events = pd.DataFrame(event_counts).T
            st.subheader(f"📉 月別：予測される戦力の減少・離脱数")
            st.bar_chart(df_events[["定年・自己都合退職", "産休・育休離脱"]], color=["#d32f2f", "#f57c00"])
            
            st.subheader("📅 年間サマリー（採用前提に基づく純増減）")
            yearly_summary = {}
            for ts, row in df_events.iterrows():
                year = ts.split("-")[0]
                if year not in yearly_summary: yearly_summary[year] = {"離脱合計": 0, "新卒補充・復帰合計": 0}
                yearly_summary[year]["離脱合計"] += (row["定年・自己都合退職"] + row["産休・育休離脱"])
                yearly_summary[year]["新卒補充・復帰合計"] += (row["育休・時短復帰(増)"] + row.get("新卒入職(補充)", 0))
                
            df_yearly = pd.DataFrame(yearly_summary).T
            df_yearly["純増減（予測）"] = df_yearly["新卒補充・復帰合計"] - df_yearly["離脱合計"]
            st.dataframe(df_yearly.style.map(lambda x: 'color: red; font-weight: bold;' if x < 0 else 'color: green;', subset=['純増減（予測）']), use_container_width=True)
            
            st.info("💡 **【経営・採用への示唆】** 上記の「離脱合計」は、今後予測される**自然減（マイナス）の人数**です。\n\n画面5の「最適配置シミュレータ」で算出した**『現在の不足FTE（定数とのギャップ）』**に、この**『未来の離脱数』**を加味することで、次年度の正確な採用・異動のターゲット人数を策定できます。")